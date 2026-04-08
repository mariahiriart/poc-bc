from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.responses import Response, FileResponse
from fastapi.staticfiles import StaticFiles
from contextlib import asynccontextmanager
import uvicorn, json, re, os, threading, urllib.request, time as _time
from datetime import datetime
import openpyxl
from dotenv import load_dotenv
load_dotenv()
import database
from typo_corrector import corregir_lista_bops

# Importar lógica de parseo desde módulo compartido
from bop_parser import (
    BO_PHONES, MEXICO_TZ,
    fmt_hour, mexico_now,
    extract_bop, is_bo_fmt, is_driver_msg,
    parse_driver, parse_bo, is_exitoso,
    get_media_for_bop,
)

# ── CONFIGURACION ──────────────────────────────────────────────────────────────
TOKEN             = os.environ.get('WHAPI_TOKEN', '')
CHAT_ID           = '120363349733984596@g.us'
JACOB_PHONE       = '5215625585843'
ROBERTO_PHONE     = '5215586931845'
XLSX_CHAT_ID      = '120363349579190170@g.us'
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

AUTHORIZED_CHAT_IDS = [
    '120363349733984596@g.us',
    '120363419653209546@g.us',
    '120363423645957323@g.us',
    '120363349579190170@g.us',
    '120363400981379542@g.us',
    '120363380129878437@g.us',
]

BO_PHONES = {'5215568660814', '5215528551646', '5215530313942', '5215580510043'}
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))

# ── ESTADO EN MEMORIA ──────────────────────────────────────────────────────────
state_lock         = threading.Lock()
bop_reports        = {}
bo_responses       = {}
last_bops_by_phone = {}
rutas_csv          = {}
bop_to_ruta        = {}
bop_por_ruta_punto = {}  # {ruta_num: {punto_int: bop}} — para corrector typos
driver_names       = {}
pending_media_by_phone = {} # buffer de evidencia reciente


MEXICO_OFFSET = -6 * 3600  # CST = UTC-6 (mantenido para compatibilidad con fmt_hour legacy)

def mexico_now_compat():
    return datetime.utcfromtimestamp(_time.time() + MEXICO_OFFSET)

# mexico_now() ya viene de bop_parser (retorna datetime aware con tzinfo)
# Para compatibilidad con código existente que usa strftime directamente:
_mexico_now_orig = mexico_now
def mexico_now():
    return _mexico_now_orig()

today_str = mexico_now().strftime('%Y-%m-%d')

# ── CARGAR XLSX DEL DIA ────────────────────────────────────────────────────────
def _parse_xlsx_bytes(raw_bytes, source_label=''):
    """
    Parsea bytes de un xlsx y devuelve dict {ruta: [bops]}.
    Centraliza la lógica para no duplicarla entre disco y Postgres.
    """
    import io
    rutas = {}
    processed = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        print(f'[XLSX] Sheet activa: {ws.title} | fuente: {source_label}', flush=True)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6: continue
            vehiculo = str(row[1]).strip() if row[1] is not None else None
            bop      = str(row[5]).strip() if row[5] is not None else None
            if not vehiculo or not bop or len(bop) < 4:
                continue
            rutas.setdefault(vehiculo, [])
            if bop not in rutas[vehiculo]:
                rutas[vehiculo].append(bop)
                processed += 1
        print(f'[XLSX] Parseado: {len(rutas)} rutas, {processed} BOPs.', flush=True)
    except Exception as e:
        print(f'[XLSX] Error parseando xlsx ({source_label}): {e}', flush=True)
    return rutas


def _build_ruta_punto_index(raw_bytes):
    """
    Construye {ruta_num: {punto_int: bop}} desde bytes de xlsx.
    Usado por el typo_corrector para la Regla 1 (ruta + punto exactos).
    row[1]=vehiculo, row[2]=parada(punto), row[5]=bop
    """
    import io, re as _re
    index = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6: continue
            vehiculo = str(row[1]).strip() if row[1] is not None else None
            punto    = row[2]
            bop      = str(row[5]).strip() if row[5] is not None else None
            if not vehiculo or punto is None or not bop or len(bop) < 4:
                continue
            m = _re.search(r'\d+', vehiculo)
            ruta_num = m.group(0) if m else vehiculo
            try:
                punto_int = int(punto)
            except (ValueError, TypeError):
                continue
            index.setdefault(ruta_num, {})[punto_int] = bop
    except Exception as e:
        print(f'[XLSX] Error construyendo indice ruta+punto: {e}', flush=True)
    return index


def load_xlsx(fecha_str):
    """
    Carga el xlsx de rutas para la fecha dada.
    Orden de prioridad:
      1. Disco (archivo con nombre exacto del día)
      2. Postgres (xlsx_files table)
      3. Disco fallback filtrado por día (evita cargar xlsx del día incorrecto)
    """
    global rutas_csv, bop_to_ruta
    import glob, io
    parts = fecha_str.split('-')
    if len(parts) < 3: return
    d       = parts[2]        # '28' (sin padding puede ser '7')
    day_int = int(d)

    # 1. Buscar en disco con patrón exacto del día
    patron    = os.path.join(BASE_DIR, f'rutas_{d}_*.xlsx')
    potential = glob.glob(patron)
    if not potential:
        # También probar con padding (rutas_07_mzo.xlsx para día 7)
        patron_pad = os.path.join(BASE_DIR, f'rutas_{day_int:02d}_mzo.xlsx')
        if os.path.exists(patron_pad):
            potential = [patron_pad]

    rutas = {}

    if potential:
        fname = potential[0]
        print(f'[XLSX] Cargando desde disco: {fname}', flush=True)
        try:
            with open(fname, 'rb') as f:
                raw = f.read()
            rutas = _parse_xlsx_bytes(raw, source_label=fname)
            _last_raw = raw
        except Exception as e:
            print(f'[XLSX] Error leyendo disco {fname}: {e}', flush=True)

    # 2. Fallback Postgres si el disco no dio resultados
    if not rutas:
        print(f'[XLSX] No encontrado en disco — buscando en Postgres para {fecha_str}...', flush=True)
        result = database.load_route_file_bytes(fecha_str)
        if result:
            filename, raw_bytes = result
            # Guardar en disco para futuros accesos
            out_path = os.path.join(BASE_DIR, f'rutas_{day_int:02d}_mzo.xlsx')
            try:
                with open(out_path, 'wb') as f:
                    f.write(raw_bytes)
                print(f'[XLSX] Restaurado desde Postgres a disco: {out_path}', flush=True)
            except Exception as e:
                print(f'[XLSX] No se pudo escribir disco (continuamos desde memoria): {e}', flush=True)
            rutas = _parse_xlsx_bytes(raw_bytes, source_label=f'Postgres:{filename}')
            _last_raw = raw_bytes

    # 3. Último recurso: cualquier xlsx del día correcto en disco (evita día incorrecto)
    if not rutas:
        all_candidates = glob.glob(os.path.join(BASE_DIR, 'rutas*.xlsx'))
        filtered = [
            f for f in all_candidates
            if re.search(rf'rutas[_\-]?0?{day_int}[_\-]', os.path.basename(f), re.I)
        ]
        if filtered:
            print(f'[XLSX] Fallback filtrado por día {day_int}: {filtered[0]}', flush=True)
            try:
                with open(filtered[0], 'rb') as f:
                    raw = f.read()
                rutas = _parse_xlsx_bytes(raw, source_label=filtered[0])
                _last_raw = raw
            except Exception as e:
                print(f'[XLSX] Error en fallback filtrado: {e}', flush=True)

    if not rutas:
        print(f'[XLSX] ADVERTENCIA: No se encontró xlsx para {fecha_str} en disco ni Postgres.', flush=True)
        return

    # Construir indice ruta+punto para el typo corrector
    # _last_raw se captura dentro del bloque que cargó las rutas exitosamente
    _ruta_punto_idx = _build_ruta_punto_index(_last_raw) if '_last_raw' in dir() and _last_raw else {}

    with state_lock:
        rutas_csv          = rutas
        bop_to_ruta        = {b: r for r, bops in rutas.items() for b in bops}
        bop_por_ruta_punto = _ruta_punto_idx
    print(f'[XLSX] SUCCESS: {len(rutas_csv)} rutas, {sum(len(v) for v in rutas_csv.values())} BOPs cargados.', flush=True)

# ── ASIGNACION DRIVER-RUTA DESDE IMAGEN DE ROBERTO ────────────────────────────
def _descargar_media(media_id):
    url     = f'https://gate.whapi.cloud/media/{media_id}'
    headers = {'Authorization': f'Bearer {TOKEN}', 'User-Agent': 'Mozilla/5.0'}
    req     = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=30) as r:
        return r.read()

def procesar_imagen_asignacion(img_id):
    global driver_names
    try:
        img_data = _descargar_media(img_id)
    except Exception as e:
        print(f'[ASIG] Error descargando imagen {img_id}: {e}', flush=True)
        return
    img_path = os.path.join(BASE_DIR, 'asignacion_hoy.jpg')
    with open(img_path, 'wb') as f:
        f.write(img_data)
    print(f'[ASIG] Imagen guardada: {img_path} ({len(img_data)} bytes)', flush=True)
    if not ANTHROPIC_API_KEY:
        print('[ASIG] Sin ANTHROPIC_API_KEY — imagen guardada pero no procesada.', flush=True)
        return
    import anthropic, base64 as _b64
    client  = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    img_b64 = _b64.standard_b64encode(img_data).decode('utf-8')
    try:
        resp = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=600,
            messages=[{'role': 'user', 'content': [
                {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': img_b64}},
                {'type': 'text', 'text':
                    'Esta tabla tiene columnas "Nombre" (numero de ruta/operador) y "Nombre Completo" (nombre real del driver). '
                    'Extrae cada fila como JSON array: [{"ruta": N, "nombre": "Nombre Completo"}]. '
                    'Solo el JSON array, sin texto ni markdown extra.'}
            ]}]
        )
        raw         = resp.content[0].text.strip()
        raw         = re.sub(r'^```[a-z]*\n?', '', raw).rstrip('`').strip()
        assignments = json.loads(raw)
        new_names   = {f'RUTA {a["ruta"]}': a['nombre'] for a in assignments if 'ruta' in a and 'nombre' in a}
        with state_lock:
            driver_names.update(new_names)
        dn_path = os.path.join(BASE_DIR, 'driver_names.json')
        with open(dn_path, 'w', encoding='utf-8') as f:
            json.dump(driver_names, f, ensure_ascii=False, indent=2)
        print(f'[ASIG] Asignaciones cargadas: {new_names}', flush=True)
        # Persistir en Postgres para sobrevivir reinicios de Render
        try:
            with state_lock:
                fecha_hoy = today_str
            database.save_driver_assignments(fecha_hoy, dict(driver_names))
        except Exception as e2:
            print(f'[ASIG] Error guardando assignments en Postgres: {e2}', flush=True)
    except Exception as e:
        print(f'[ASIG] Error procesando con Claude Vision: {e}', flush=True)

def cargar_driver_names_desde_disco():
    """
    Carga driver names desde disco (driver_names.json).
    Si el archivo no existe, intenta recuperar desde Postgres (driver_assignments).
    """
    global driver_names
    dn_path = os.path.join(BASE_DIR, 'driver_names.json')
    if os.path.exists(dn_path):
        try:
            with open(dn_path, encoding='utf-8') as f:
                driver_names = json.load(f)
            print(f'[ASIG] Driver names cargados desde disco: {len(driver_names)} rutas', flush=True)
            return  # disco OK, no necesitamos Postgres
        except Exception as e:
            print(f'[ASIG] Error leyendo driver_names.json: {e}', flush=True)

    # Fallback Postgres
    with state_lock:
        fecha = today_str
    try:
        pg_names = database.load_driver_assignments(fecha)
        if pg_names:
            with state_lock:
                driver_names = pg_names
            # Restaurar en disco para futuros accesos
            try:
                with open(dn_path, 'w', encoding='utf-8') as f:
                    json.dump(driver_names, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
            print(f'[ASIG] Driver names recuperados desde Postgres: {len(driver_names)} rutas', flush=True)
        else:
            print(f'[ASIG] Sin driver names en disco ni Postgres para {fecha}.', flush=True)
    except Exception as e:
        print(f'[ASIG] Error cargando driver names desde Postgres: {e}', flush=True)

# ── DESCARGA AUTOMATICA DE XLSX DESDE CUALQUIER CHAT AUTORIZADO ────────────────
def descargar_xlsx_doc(doc_id, filename):
    # FIX: Regex corregido — soporta espacios, guiones bajos o sin separador.
    # Antes: r'(\d+)\s+(MZO|marzo)' fallaba con nombres tipo "rutas_27_mzo.xlsx"
    m = re.search(r'(\d+)[\s_]*(MZO|marzo)', filename, re.I)
    day = int(m.group(1)) if m else mexico_now().day
    out_name = f'rutas_{day:02d}_mzo.xlsx'
    out_path = os.path.join(BASE_DIR, out_name)
    try:
        data = _descargar_media(doc_id)
        with open(out_path, 'wb') as f:
            f.write(data)
        print(f'[XLSX] Descargado: {out_name} ({len(data)} bytes) desde {filename}', flush=True)
        # Persistir en Postgres para sobrevivir reinicios de Render
        # Persistir en Postgres y poblar routes + route_stops
        try:
            from datetime import date as _date
            y_str, m_str = mexico_now().strftime('%Y-%m').split('-')
            fecha_xlsx = f'{y_str}-{m_str}-{day:02d}'
            database.save_route_file(fecha_xlsx, out_name, data)
            database.import_routes_from_xlsx(fecha_xlsx, data, out_name)
        except Exception as e2:
            print(f'[XLSX] Error guardando xlsx en Postgres: {e2}', flush=True)
        # Recargar inmediatamente las rutas del día actual si el archivo coincide
        hoy = mexico_now()
        if day == hoy.day:
            load_xlsx(hoy.strftime('%Y-%m-%d'))
            print(f'[XLSX] Dashboard actualizado con {len(rutas_csv)} rutas y {sum(len(v) for v in rutas_csv.values())} BOPs.', flush=True)
    except Exception as e:
        print(f'[XLSX] Error descargando {doc_id}: {e}', flush=True)

# ── CONSTRUIR PAYLOAD DEL DASHBOARD ──────────────────────────────────────────
def _build_payload_from_state(reps, bos, rutas, b2r, dn, fecha):
    """
    Núcleo de construcción del payload del dashboard.
    Recibe estado como dicts puros — sin tocar globales.
    Usado tanto para el día en vivo (desde RAM) como para históricos (desde DB).
    """
    detalle           = []
    detalle_extras    = []   # IDs reportados que NO están en el Excel
    all_bops          = set()
    excel_bops        = set(b2r.keys())  # Todos los IDs del Excel
    tiene_rutas       = bool(b2r)

    # ── Reportes reconocidos (presentes en el Excel) ─────────────────────────
    for bop, rep in reps.items():
        bo       = bos.get(bop, {})
        ruta_eff = b2r.get(bop) or rep.get('ruta') or '?'
        exitoso  = is_exitoso(rep.get('status', ''))

        item = {
            'bop':           bop,
            'driver':        dn.get(ruta_eff) or rep.get('nombre', ''),
            'ruta':          ruta_eff,
            'status_final':  'Exito' if exitoso else 'Fallido / Incidencia',
            'evidencias':    rep.get('imgs', 0),
            'media':         rep.get('media', []),
            'driver_status': rep.get('status') or 'sin estatus',
            'driver_obs':    rep.get('obs', ''),
            'bo_status':     bo.get('bo_status', 'N/A'),
            'bo_obs':        bo.get('bo_obs', ''),
            'raw_drv_msgs':  rep.get('msgs', []),
            'raw_bo_msgs':   bo.get('msgs', []),
            'ultima_hora':   rep.get('hora', ''),
        }

        if tiene_rutas and bop not in excel_bops:
            # Este BOP fue reportado pero NO está en el Excel → extra/no reconocido
            item['no_reconocido'] = True
            detalle_extras.append(item)
        else:
            detalle.append(item)
            all_bops.add(bop)

    # ── Solo BO, sin reporte de driver ───────────────────────────────────────
    for bop, bo in bos.items():
        if bop in all_bops:
            continue
        ruta_eff = b2r.get(bop, '?')
        item = {
            'bop': bop, 'driver': dn.get(ruta_eff) or 'Solo BO',
            'ruta': ruta_eff,
            'status_final': 'Fallido / Incidencia',
            'evidencias': 0, 'media': [],
            'driver_status': 'N/A', 'driver_obs': '',
            'bo_status':    bo.get('bo_status', 'N/A'),
            'bo_obs':       bo.get('bo_obs', ''),
            'raw_drv_msgs': [], 'raw_bo_msgs': bo.get('msgs', []),
            'ultima_hora':  bo.get('hora', ''),
        }
        if tiene_rutas and bop not in excel_bops:
            item['no_reconocido'] = True
            detalle_extras.append(item)
        else:
            detalle.append(item)
            all_bops.add(bop)

    # ── Sin reporte: asignados pero sin ningún mensaje ────────────────────────
    sin_reporte_vistos = set()
    for ruta, bops_asig in rutas.items():
        for bop in bops_asig:
            if bop not in all_bops and bop not in sin_reporte_vistos:
                detalle.append({
                    'bop': bop, 'driver': dn.get(ruta, '—'), 'ruta': ruta,
                    'status_final': 'Sin Reporte',
                    'evidencias': 0, 'media': [],
                    'driver_status': 'Sin reporte', 'driver_obs': '',
                    'bo_status': 'N/A', 'bo_obs': '',
                    'raw_drv_msgs': [], 'raw_bo_msgs': [],
                    'ultima_hora': '—',
                })
                sin_reporte_vistos.add(bop)

    detalle.sort(key=lambda x: (x['ruta'], x['bop']))
    # Los extras van al final, ordenados también
    detalle_extras.sort(key=lambda x: x['bop'])

    total    = len(detalle)            # Solo Excel BOPs — no cambia con extras
    exitosos = sum(1 for d in detalle if d['status_final'] == 'Exito')
    fallidos = sum(1 for d in detalle if d['status_final'] == 'Fallido / Incidencia')
    con_bo   = sum(1 for d in detalle if d['bo_status'] != 'N/A')
    faltantes = total - (exitosos + fallidos)
    no_reconocidos = len(detalle_extras)

    # Combinamos al final: primero los reconocidos, luego los extras
    detalle = detalle + detalle_extras

    rutas_out = []
    for ruta, bops_asig in sorted(rutas.items(),
            key=lambda x: int(x[0].split()[-1]) if len(x[0].split()) > 1 and x[0].split()[-1].isdigit() else 0):
        reps_bops = [b for b in bops_asig if b in all_bops]
        pend_bops = [b for b in bops_asig if b not in all_bops]
        driver = dn.get(ruta) or next(
            (d['driver'] for d in detalle
             if d['ruta'] == ruta and d['driver'] not in ('Solo BO', '—')), '?')
        rutas_out.append({
            'ruta': ruta, 'driver': driver,
            'total_asignado':  len(bops_asig),
            'total_reportado': len(reps_bops),
            'total_faltante':  len(pend_bops),
            'faltantes_bops':  pend_bops,
        })

    return {
        'generado_at': mexico_now().strftime('%Y-%m-%d %H:%M:%S'),
        'fecha':       fecha,
        'kpis': {
            'total_asignados':   total,
            'total_reportados':  exitosos + fallidos,
            'total_sin_reporte': faltantes,
            'exitosos':          exitosos,
            'fallidos':          fallidos,
            'con_bo':            con_bo,
            'pct_exito':         round(exitosos / total * 100, 1) if total else 0,
            'no_reconocidos':    no_reconocidos,
        },
        'rutas':              rutas_out,
        'detalle_reportados': detalle,
    }


def _build_dashboard_payload():
    """Payload del dashboard desde RAM (día en vivo). Sin I/O."""
    with state_lock:
        return _build_payload_from_state(
            dict(bop_reports),
            dict(bo_responses),
            dict(rutas_csv),
            dict(bop_to_ruta),
            dict(driver_names),
            today_str,
        )

def regenerar_dashboard():
    """Solo actualiza KPIs desde memoria. Ya NO escribe ningun archivo local."""
    payload = _build_dashboard_payload()
    return payload['kpis'], len(payload['detalle_reportados'])

# ── PROCESAR UN MENSAJE ────────────────────────────────────────────────────────
def _add_media_to_bop(bop, item):
    if bop and bop in bop_reports:
        media_list = bop_reports[bop].setdefault('media', [])
        # Evitar duplicados exactos
        if item.get('id'):
            if any(m.get('id') == item['id'] for m in media_list): return
        elif item.get('type') == 'location':
            if any(m.get('type') == 'location' and m.get('lat') == item['lat'] and m.get('lon') == item['lon'] for m in media_list): return

        media_list.append(item)
        if item['type'] == 'image':
            bop_reports[bop]['imgs'] = bop_reports[bop].get('imgs', 0) + 1

def _buffer_media_and_assign(phone, item, target_bops, ts):
    # Guarda en el buffer (hasta 10 min) y asigna a target_bops actual
    with state_lock:
        queue = pending_media_by_phone.setdefault(phone, [])
        queue.append({'ts': ts, 'item': item})
        # Limpiar
        pending_media_by_phone[phone] = [x for x in queue if ts - x['ts'] < 600]

        if target_bops:
            for bop in target_bops:
                _add_media_to_bop(bop, item)

def _retro_assign_buffered_media(phone, new_bops, ts):
    # Re-asigna media de los ultimos 5 min a estos nuevos bops (por si llegaron antes del texto)
    with state_lock:
        queue = pending_media_by_phone.get(phone, [])
        for q in queue:
            if abs(ts - q['ts']) < 300: # ventana de 5 mins
                for bop in new_bops:
                    _add_media_to_bop(bop, q['item'])

def procesar_mensaje(msg):
    """
    Parsea el mensaje y actualiza la RAM.
    Devuelve un dict describiendo lo que se parseó (para que el caller
    lo persista en PostgreSQL), o None si el mensaje no era un reporte.
    """
    phone  = msg.get('from', '')
    nombre = msg.get('from_name') or phone
    ts     = msg.get('timestamp', 0)
    mtype  = msg.get('type', '')
    hora   = fmt_hour(ts) if ts else ''

    if mtype == 'location':
        loc = msg.get('location', {})
        lat = loc.get('latitude') or loc.get('lat')
        lon = loc.get('longitude') or loc.get('lng')
        if lat and lon:
            item = {
                'type':    'location',
                'preview': loc.get('preview', ''),
                'url':     f'https://maps.google.com/?q={float(lat)},{float(lon)}',
                'lat':     float(lat), 'lon': float(lon),
            }
            with state_lock:
                last_bops = last_bops_by_phone.get(phone, [])
            _buffer_media_and_assign(phone, item, last_bops, ts)
        return None

    if mtype == 'image':
        img     = msg.get('image') or {}
        caption = img.get('caption', '') or ''
        item    = {
            'type':    'image',
            'preview': img.get('preview', ''),
            'id':      img.get('id', ''),
            'caption': caption,
        }
        caption_bops = extract_bop(caption) if caption else []
        with state_lock:
            target_bops = caption_bops if caption_bops else last_bops_by_phone.get(phone, [])

        # Fallback DB si no hay contexto en memoria
        if not target_bops:
            try:
                last_db_bops = database.get_last_bops_for_phone(phone, today_str)
                if last_db_bops:
                    target_bops = last_db_bops
                    print(f'[MEDIA] Asociando imagen a BOPs del DB: {target_bops}', flush=True)
            except Exception: pass

        if target_bops:
            pass # Solo para logica; la asignacion real se hace abajo con _buffer_media

        _buffer_media_and_assign(phone, item, target_bops, ts)

        if not caption:
            return None
        text = caption

    if mtype == 'video':
        vid     = msg.get('video') or {}
        caption = vid.get('caption', '') or ''
        item    = {
            'type':    'video',
            'preview': vid.get('preview', ''),
            'id':      vid.get('id', ''),
            'caption': caption,
        }
        caption_bops = extract_bop(caption) if caption else []
        with state_lock:
            target_bops = caption_bops if caption_bops else last_bops_by_phone.get(phone, [])

        if not target_bops:
            try:
                last_db_bops = database.get_last_bops_for_phone(phone, today_str)
                if last_db_bops:
                    target_bops = last_db_bops
                    print(f'[MEDIA] Asociando video a BOPs del DB: {target_bops}', flush=True)
            except Exception: pass

        if target_bops:
            pass

        _buffer_media_and_assign(phone, item, target_bops, ts)

        if not caption:
            return None
        text = caption
    elif mtype == 'text':
        text = (msg.get('text') or {}).get('body', '')
        if not text:
            return None

        # Buscar links de Google Maps en el texto
        maps_match = re.search(r'https?://(?:maps\.google\.com|goo\.gl/maps|maps\.app\.goo\.gl)/[^\s]+', text)
        if maps_match:
            item = {
                'type': 'location',
                'url':  maps_match.group(0),
                'text': text[:50] + '...' if len(text) > 50 else text
            }
            with state_lock:
                last_bops = last_bops_by_phone.get(phone, [])
                for bop in last_bops:
                    _add_media_to_bop(bop, item)
    else:
        return None

    # ── BO ─────────────────────────────────────────────────────────────────────
    if phone in BO_PHONES or is_bo_fmt(text):
        r = parse_bo(text)
        if r:
            with state_lock:
                for bop in r['bops']:
                    if bop not in bo_responses:
                        bo_responses[bop] = {
                            'bo_status': r['bo_status'],
                            'bo_obs':    r['bo_obs'],
                            'msgs':      [],
                            'hora':      hora,
                        }
                    else:
                        bo_responses[bop]['bo_status'] = r['bo_status']
                        bo_responses[bop]['bo_obs']     = r['bo_obs']
                        bo_responses[bop]['hora']        = hora
                    bo_responses[bop]['msgs'].append(f'{hora} {nombre}: {text[:80]}')
            print(f'[RT] BO  BOPs={r["bops"]} status={r["bo_status"]}', flush=True)
            # Devolver info para persistir en DB
            return {'type': 'bo', 'bops': r['bops'], 'parsed': r}
        return None

    # ── Driver ─────────────────────────────────────────────────────────────────
    is_drv = is_driver_msg(text)
    if not is_drv:
        return None

    r = parse_driver(text)
    if r:
        all_bops_msg = r['bops']

        # ── TYPO CORRECTOR: corregir BOPs mal escritos antes de procesar ──────
        with state_lock:
            _rutas_snap = dict(rutas_csv)
            _b2r_snap   = dict(bop_to_ruta)
            _brp_snap   = dict(bop_por_ruta_punto)
        all_bops_msg, _correcciones = corregir_lista_bops(
            all_bops_msg, r.get('ruta'), r.get('punto'),
            _rutas_snap, _b2r_snap, _brp_snap,
        )
        for _c in _correcciones:
            print(
                f'[TYPO] Auto-corrección: {_c["bop_original"]} → {_c["bop_corregido"]} '
                f'(ruta={_c["ruta_real"]}, dist={_c["distancia"]}, conf={_c["confianza"]})',
                flush=True
            )
        # ─────────────────────────────────────────────────────────────────────

        with state_lock:
            for bop in all_bops_msg:
                ruta_real = bop_to_ruta.get(bop) or r['ruta'] or '?'
                if bop not in bop_reports:
                    bop_reports[bop] = {
                        'phone': phone, 'nombre': nombre, 'ruta': ruta_real,
                        'punto': r['punto'], 'status': r['status'], 'obs': r['obs'],
                        'ts': ts, 'hora': hora, 'msgs': [], 'imgs': 0, 'media': [],
                    }
                else:
                    bop_reports[bop]['status'] = r['status']
                    bop_reports[bop]['obs']    = r['obs']
                    bop_reports[bop]['hora']   = hora
                bop_reports[bop]['msgs'].append(f'{hora} {nombre}: {text[:100]}')
            last_bops_by_phone[phone] = list(all_bops_msg)

        # Rescate de media enviada ANTES del texto!
        _retro_assign_buffered_media(phone, all_bops_msg, ts)

        print(f'[RT] DRV BOPs={all_bops_msg} status={r["status"]}', flush=True)
        return {'type': 'driver', 'bops': all_bops_msg, 'parsed': r, 'nombre': nombre}
    return None

# ── ROLLOVER DE MEDIANOCHE ─────────────────────────────────────────────────────
def _reset_estado_dia():
    global bop_reports, bo_responses, last_bops_by_phone, rutas_csv, bop_to_ruta, bop_por_ruta_punto, pending_media_by_phone
    with state_lock:
        bop_reports        = {}
        bo_responses       = {}
        last_bops_by_phone = {}
        rutas_csv          = {}
        bop_to_ruta        = {}
        bop_por_ruta_punto = {}
        pending_media_by_phone = {}
    print('[ROLLOVER] Estado en memoria limpiado.', flush=True)

def _watcher_medianoche():
    global today_str
    print('[WATCHER] Hilo de medianoche activo.', flush=True)
    while True:
        _time.sleep(30)
        nuevo_dia = mexico_now().strftime('%Y-%m-%d')
        if nuevo_dia == today_str:
            continue
        print(f'[ROLLOVER] Cambio de dia: {today_str} -> {nuevo_dia}', flush=True)
        _reset_estado_dia()
        with state_lock:
            today_str = nuevo_dia
        threading.Thread(target=init_today, daemon=True).start()

# ── INICIALIZACION ─────────────────────────────────────────────────────────────
def _descargar_mensajes_whapi(fecha_str):
    y, mo, d        = int(fecha_str[:4]), int(fecha_str[5:7]), int(fecha_str[8:10])
    MEXICO_OFFSET_S = 6 * 3600
    d_start = int(datetime(y, mo, d,  0,  0,  0).timestamp()) + MEXICO_OFFSET_S
    d_end   = int(datetime(y, mo, d, 23, 59, 59).timestamp()) + MEXICO_OFFSET_S
    headers = {
        'Authorization': f'Bearer {TOKEN}',
        'Accept':        'application/json',
        'User-Agent':    'Mozilla/5.0',
    }
    all_msgs = []
    for chat_id in AUTHORIZED_CHAT_IDS:
        offset    = 0
        oldest    = 9_999_999_999
        msgs_chat = 0
        while oldest > d_start:
            url = f'https://gate.whapi.cloud/messages/list/{chat_id}?count=100&offset={offset}'
            try:
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=15) as r:
                    data = json.loads(r.read())
            except Exception as e:
                print(f'[API] Error Whapi chat={chat_id}: {e}', flush=True)
                break
            msgs = data.get('messages', [])
            if not msgs:
                break
            for m in msgs:
                ts_m = m.get('timestamp', 0)
                if d_start <= ts_m <= d_end:
                    if 'chat_id' not in m:
                        m['chat_id'] = chat_id
                    all_msgs.append(m)
                    msgs_chat += 1
                oldest = min(oldest, ts_m)
            if len(msgs) < 100:
                break
            offset += 100
        print(f'[API] Whapi {chat_id}: {msgs_chat} msgs del dia', flush=True)
    all_msgs.sort(key=lambda m: m.get('timestamp', 0))
    return all_msgs

def rescatar_archivos_contexto(fecha_str):
    import datetime as dt_lib
    y, mo, d = int(fecha_str[:4]), int(fecha_str[5:7]), int(fecha_str[8:10])
    dt_hoy = datetime(y, mo, d, 0, 0, 0)
    MEXICO_OFFSET_S = 6 * 3600

    headers = {'Authorization': f'Bearer {TOKEN}', 'Accept': 'application/json', 'User-Agent': 'Mozilla/5.0'}

    # 1. Rescatar XLSX (Ventana: 3 días atrás 16:00 hasta fin del día de hoy)
    # Cubre: xlsx enviado la noche anterior O el mismo día por la mañana
    target_fname = f'rutas_{d:02d}_mzo.xlsx'
    if not os.path.exists(os.path.join(BASE_DIR, target_fname)):
        dt_ayer_18 = dt_hoy - dt_lib.timedelta(hours=32)  # 3 días atrás 16:00 cubre sábado
        dt_fin_xlsx = dt_hoy + dt_lib.timedelta(hours=24) # Fin del día de hoy (cubre mañana)
        ts_inicio = int(dt_ayer_18.timestamp()) + MEXICO_OFFSET_S
        ts_fin    = int(dt_fin_xlsx.timestamp()) + MEXICO_OFFSET_S

        try:
            url = f'https://gate.whapi.cloud/messages/list/{XLSX_CHAT_ID}?count=100'
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=15) as r:
                data = json.loads(r.read())
            for m in data.get('messages', []):
                ts_m = m.get('timestamp', 0)
                if ts_inicio <= ts_m <= ts_fin and m.get('type') == 'document':
                    doc = m.get('document') or {}
                    fname = doc.get('filename', '')
                    if fname.lower().endswith('.xlsx') and doc.get('id'):
                        print(f'[INIT] Rescatando XLSX perdido de anoche: {fname}', flush=True)
                        descargar_xlsx_doc(doc.get('id'), fname)
                        break
        except Exception as e:
            print(f'[INIT] Error rescatando XLSX: {e}')

    # Ventana de imagen de asignación: desde 20:00 de ayer hasta medianoche de hoy.
    # Cubre madrugadas, reinicios tardíos de Render y cualquier hora del día.
    dt_inicio_img = dt_hoy - dt_lib.timedelta(hours=4)   # 20:00 hora Mexico de ayer
    dt_fin_img    = dt_hoy + dt_lib.timedelta(hours=24)  # Medianoche de hoy (día completo)
    ts_inicio_img = int(dt_inicio_img.timestamp()) + MEXICO_OFFSET_S
    ts_fin_img    = int(dt_fin_img.timestamp()) + MEXICO_OFFSET_S

    try:
        url = f'https://gate.whapi.cloud/messages/list/{XLSX_CHAT_ID}?count=100'
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read())
        for m in data.get('messages', []):
            ts_m = m.get('timestamp', 0)
            if ts_inicio_img <= ts_m <= ts_fin_img and m.get('type') == 'image' and m.get('from') == ROBERTO_PHONE:
                img_id = (m.get('image') or {}).get('id', '')
                if img_id:
                    print(f'[INIT] Rescatando imagen asignacion de Roberto (ventana ampliada): {img_id}', flush=True)
                    threading.Thread(target=procesar_imagen_asignacion, args=(img_id,), daemon=True).start()
                    break
    except Exception as e:
        print(f'[INIT] Error rescatando Imagen: {e}')

def init_today():
    global today_str
    fecha = mexico_now().strftime('%Y-%m-%d')
    with state_lock:
        today_str = fecha
    print(f'[API] Inicializando dia {fecha}...', flush=True)

    # Rescatar archivos de contexto (Excel e Imagen) en sus ventanas de tiempo
    if TOKEN:
        rescatar_archivos_contexto(fecha)

    load_xlsx(fecha)
    cargar_driver_names_desde_disco()

    msgs_hoy = []
    try:
        msgs_hoy = database.load_day_messages(fecha)
        print(f'[API] {len(msgs_hoy)} mensajes cargados desde BD.', flush=True)
    except Exception as e:
        print(f'[API] Error cargando BD: {e}', flush=True)

    if not msgs_hoy and TOKEN:
        print(f'[API] BD vacia — descargando desde Whapi...', flush=True)
        msgs_hoy = _descargar_mensajes_whapi(fecha)

    for m in msgs_hoy:
        procesar_mensaje(m)

    kpis, total = regenerar_dashboard()
    print(f'[API] Listo: {total} BOPs | exitosos={kpis["exitosos"]} fallidos={kpis["fallidos"]}', flush=True)

# ── FASTAPI APP ────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app):
    # Crear tablas auxiliares en Postgres si no existen (xlsx_files, driver_assignments)
    try:
        database.ensure_tables()
    except Exception as e:
        print(f'[STARTUP] ensure_tables error (no crítico): {e}', flush=True)
    threading.Thread(target=init_today,          daemon=True).start()
    threading.Thread(target=_watcher_medianoche, daemon=True).start()
    yield

app = FastAPI(title='Silent Listener RT - Brightcell/JCR v3.0', lifespan=lifespan)

# Dashboards historicos estaticos
for _d, _slug in [
    ('dashboard_2026_03_18', 'dashboard18'),
    ('dashboard_2026_03_19', 'dashboard19'),
    ('dashboard_2026_03_20', 'dashboard20'),
    ('dashboard_2026_03_21', 'dashboard21'),
    ('dashboard_2026_03_23', 'dashboard23'),
]:
    _p = os.path.join(BASE_DIR, _d)
    if os.path.isdir(_p):
        app.mount(f'/{_slug}', StaticFiles(directory=_p, html=True), name=_slug)

# Dashboard principal: historial completo (dias 18-24 + hoy)
from fastapi.responses import RedirectResponse

@app.get('/dashboard', include_in_schema=False)
def dashboard_redirect():
    return RedirectResponse(url='/dashboard/', status_code=301)

@app.get('/dashboard/', include_in_schema=False)
def dashboard_root():
    filepath = os.path.join(BASE_DIR, 'dashboard', 'index_dashboard.html')
    if os.path.exists(filepath):
        return FileResponse(filepath)
    raise HTTPException(status_code=503, detail='dashboard/index_dashboard.html no encontrado')

@app.get('/dashboard/{filename:path}', include_in_schema=False)
def dashboard_file(filename: str):
    filepath = os.path.join(BASE_DIR, 'dashboard', filename)
    if os.path.exists(filepath):
        return FileResponse(filepath)
    raise HTTPException(status_code=404, detail=f'Archivo no encontrado: {filename}')

@app.get('/')
def health():
    with state_lock:
        total    = len(bop_reports) + len(set(bo_responses) - set(bop_reports))
        exitosos = sum(1 for r in bop_reports.values() if is_exitoso(r['status']))
    return {'status': 'online', 'version': '3.2', 'fecha': today_str,
            'bops_vivos': total, 'exitosos': exitosos}

@app.get('/media/{media_id:path}')
def proxy_media(media_id: str):
    if not TOKEN:
        raise HTTPException(status_code=503, detail='Sin WHAPI token')
    url     = f'https://gate.whapi.cloud/media/{media_id}'
    headers = {'Authorization': f'Bearer {TOKEN}', 'User-Agent': 'Mozilla/5.0'}
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=20) as r:
            data         = r.read()
            content_type = r.headers.get('Content-Type', 'image/jpeg')
        return Response(content=data, media_type=content_type,
                        headers={'Cache-Control': 'public, max-age=86400'})
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))

@app.get('/api/asignacion-imagen')
def get_asignacion_imagen():
    for candidate in ['asignacion_hoy.jpg', 'asignacion_roberto.jpg']:
        img_path = os.path.join(BASE_DIR, candidate)
        if os.path.exists(img_path):
            return FileResponse(img_path, media_type='image/jpeg', headers={'Cache-Control': 'no-cache'})
    raise HTTPException(status_code=404, detail='No hay imagen de asignación para hoy')

@app.get('/status')
def get_status():
    kpis, total = regenerar_dashboard()
    return {'ok': True, 'kpis': kpis, 'total_bops': total}

@app.get('/api/reload-names')
def api_reload_names():
    cargar_driver_names_desde_disco()
    return {'ok': True, 'count': len(driver_names), 'names': driver_names}

@app.post('/api/reload-routes')
def reload_routes():
    """Recarga el xlsx de rutas del día actual desde disco sin reiniciar el servidor."""
    init_today()
    return {'ok': True, 'msg': 'Rutas y mensajes de hoy recargados correctamente'}

@app.get('/api/debug-files')
def debug_files():
    return {
        'cwd': os.getcwd(),
        'base_dir': BASE_DIR,
        'root_files': os.listdir('.'),
        'base_files': os.listdir(BASE_DIR)
    }

@app.post('/api/backfill')
def api_backfill():
    import backfill_today
    backfill_today.backfill_today()
    init_today()
    return {'ok': True}

@app.post('/api/reload-messages')
def reload_messages():
    """Limpia el estado del dia y recarga mensajes desde la BD (util tras fetch_current.py)."""
    with state_lock:
        fecha = today_str
    _reset_estado_dia()
    load_xlsx(fecha)
    msgs = []
    try:
        msgs = database.load_day_messages(fecha)
        print(f'[RELOAD] {len(msgs)} mensajes cargados desde BD para {fecha}.', flush=True)
    except Exception as e:
        print(f'[RELOAD] Error cargando BD: {e}', flush=True)
    for m in msgs:
        procesar_mensaje(m)
    kpis, total = regenerar_dashboard()
    return {'ok': True, 'fecha': fecha, 'mensajes': len(msgs), 'bops': total, 'kpis': kpis}

@app.get('/api/data')
def api_data():
    """Fuente de verdad del dashboard — solo hoy (en vivo)."""
    return _build_dashboard_payload()

# ── NUEVO ENDPOINT: Subida manual de xlsx para cualquier fecha ─────────────────
@app.post('/api/upload-xlsx/{fecha}')
async def upload_xlsx_for_date(fecha: str, file: UploadFile = File(...)):
    """
    Sube manualmente el xlsx de rutas para una fecha dada (YYYY-MM-DD).
    Corrige el estado del dashboard si el xlsx se guardó con nombre incorrecto.

    Uso (bash):
        curl -X POST https://poc-bc.onrender.com/api/upload-xlsx/2026-03-27 \\
             -F "file=@rutas_27_mzo.xlsx"

        curl -X POST https://poc-bc.onrender.com/api/upload-xlsx/2026-03-28 \\
             -F "file=@Rutas_para_28_Mzo_Telcel_Movistar_2025_JCR.xlsx"
    """
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', fecha):
        raise HTTPException(400, 'Formato de fecha inválido — usar YYYY-MM-DD')

    day      = int(fecha.split('-')[2])
    out_name = f'rutas_{day:02d}_mzo.xlsx'
    out_path = os.path.join(BASE_DIR, out_name)

    content = await file.read()
    with open(out_path, 'wb') as f:
        f.write(content)
    print(f'[XLSX] Upload manual: {out_name} ({len(content)} bytes)', flush=True)
    # Guardar en Postgres y poblar routes + route_stops
    try:
        database.save_route_file(fecha, out_name, content)
        database.import_routes_from_xlsx(fecha, content, out_name)
    except Exception as e:
        print(f'[XLSX] Error guardando upload en Postgres: {e}', flush=True)

    with state_lock:
        fecha_hoy = today_str

    if fecha == fecha_hoy:
        # Es el día actual: recargar estado en vivo
        load_xlsx(fecha)
        with state_lock:
            rc = len(rutas_csv)
            bc = sum(len(v) for v in rutas_csv.values())
        return {
            'ok': True, 'archivo': out_name,
            'rutas': rc, 'bops': bc,
            'recargado_en_vivo': True,
            'siguiente_paso': 'Dashboard actualizado. Recargá la página.'
        }

    return {
        'ok': True, 'archivo': out_name,
        'recargado_en_vivo': False,
        'siguiente_paso': f'Archivo guardado. El histórico del {fecha} se reconstruye automáticamente desde /api/history.'
    }

# ── NUEVO ENDPOINT: Forzar rescate de imagen de asignación ────────────────────
@app.post('/api/rescatar-asignacion')
def rescatar_asignacion():
    """
    Vuelve a buscar y procesar la imagen de asignación de Roberto
    en la ventana ampliada (20:00 ayer - 09:00 hoy).
    Útil cuando la imagen se mandó de madrugada y no fue capturada al inicio.
    """
    if not TOKEN:
        raise HTTPException(503, 'Sin WHAPI token configurado')
    with state_lock:
        fecha = today_str
    threading.Thread(target=rescatar_archivos_contexto, args=(fecha,), daemon=True).start()
    return {'ok': True, 'msg': f'Rescate iniciado para {fecha}. Revisá los logs y recargá /api/reload-names en ~15s.'}


@app.get('/api/thread/{bop_id}')
def api_thread(bop_id: str, fecha: str = None):
    """
    Devuelve el hilo cronológico completo de un BOP:
    todos los mensajes (texto, imágenes, ubicaciones) del driver
    y del backoffice, ordenados por timestamp real desde PostgreSQL.

    Query params:
        fecha: 'YYYY-MM-DD' en hora México. Default = hoy.
    """
    bop_id = bop_id.strip()
    if not re.match(r'^\d{7}$', bop_id):
        raise HTTPException(status_code=400, detail='BOP ID debe tener exactamente 7 dígitos')

    with state_lock:
        fecha_local = fecha or today_str

    try:
        thread = database.load_bop_thread(bop_id, fecha_local)

        # INYECCIÓN DE MEDIA HUÉRFANA DESDE LA MEMORIA
        with state_lock:
            mem_media = (bop_reports.get(bop_id) or {}).get('media', [])

        thread_ids  = {t.get('id') for t in thread if t.get('id')}
        thread_locs = {f"{t['location']['lat']}_{t['location']['lon']}" for t in thread if t.get('location')}

        for m in mem_media:
            if m.get('id') and m['id'] in thread_ids: continue
            if m.get('type') == 'location' and f"{m.get('lat')}_{m.get('lon')}" in thread_locs: continue

            dummy_item = {
                'id': m.get('id', ''),
                'sender_name': 'Driver (Evidencia)',
                'sender_phone': '',
                'type': m.get('type') or 'image',
                'sent_at': '',
                'hora': m.get('hora', ''),
                'is_bo': False,
                'content': '',
                'media': None,
                'location': None
            }
            if m.get('type') == 'location':
                dummy_item['location'] = {
                    'lat': m.get('lat'), 'lon': m.get('lon'),
                    'url': m.get('url')
                }
            else:
                dummy_item['media'] = m

            thread.append(dummy_item)

        def get_hora(t):
            return t.get('hora') or '99:99'
        thread.sort(key=get_hora)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    with state_lock:
        ruta   = bop_to_ruta.get(bop_id, '?')
        dn_now = dict(driver_names)

    return {
        'bop_id': bop_id,
        'fecha':  fecha_local,
        'ruta':   ruta,
        'driver': dn_now.get(ruta, '?'),
        'thread': thread,
        'total':  len(thread),
    }


def _build_day_payload_from_db(fecha_str):
    """
    Reconstruye el payload del dashboard para un día histórico leyendo
    mensajes desde PostgreSQL. No toca el estado global en absoluto.
    Usa _build_payload_from_state() — misma lógica que el día en vivo.
    """
    # 1. Rutas: primero xlsx del dia (con nombre exacto o con padding), luego DB
    day        = fecha_str.split('-')[2]          # '27' (sin padding)
    day_padded = f'{int(day):02d}'                # '27' (ya tiene 2 dígitos)

    # Buscar el xlsx con ambas variantes de nombre
    fname_xlsx = None
    for candidate in [
        os.path.join(BASE_DIR, f'rutas_{day}_mzo.xlsx'),
        os.path.join(BASE_DIR, f'rutas_{day_padded}_mzo.xlsx'),
    ]:
        if os.path.exists(candidate):
            fname_xlsx = candidate
            break

    local_rutas = {}
    if fname_xlsx:
        try:
            # FIX: data_only=True para obtener valores calculados, no fórmulas
            wb = openpyxl.load_workbook(fname_xlsx, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # FIX: str() en vehiculo para evitar crash si la celda es numérica
                vehiculo = str(row[1]).strip() if row[1] is not None else None
                bop      = str(row[5]).strip() if row[5] is not None else None
                # FIX: mismo criterio que load_xlsx (len < 4) en lugar de != 7
                if not vehiculo or not bop or len(bop) < 4:
                    continue
                ruta = vehiculo
                local_rutas.setdefault(ruta, [])
                if bop not in local_rutas[ruta]:
                    local_rutas[ruta].append(bop)
            print(f'[HIST] {fecha_str}: xlsx cargado desde {fname_xlsx} — {sum(len(v) for v in local_rutas.values())} BOPs', flush=True)
        except Exception as e:
            print(f'[HIST] Error leyendo xlsx {fname_xlsx}: {e}')

    if not local_rutas:
        # Intentar recuperar xlsx desde Postgres antes de ir a routes/route_stops
        try:
            pg_result = database.load_route_file_bytes(fecha_str)
            if pg_result:
                _, raw_bytes = pg_result
                local_rutas = _parse_xlsx_bytes(raw_bytes, source_label=f'Postgres histórico {fecha_str}')
                print(f'[HIST] {fecha_str}: rutas desde xlsx en Postgres — {sum(len(v) for v in local_rutas.values())} BOPs', flush=True)
        except Exception as e:
            print(f'[HIST] Error cargando xlsx desde Postgres: {e}', flush=True)

    if not local_rutas:
        local_rutas, _ = database.load_routes_from_db(fecha_str)
        print(f'[HIST] {fecha_str}: rutas cargadas desde routes/route_stops DB — {len(local_rutas)} rutas', flush=True)

    local_bop_to_ruta = {b: r for r, bops in local_rutas.items() for b in bops}

    # 2. Driver names del dia
    local_dn = {}
    for dn_path in (
        os.path.join(BASE_DIR, f'driver_names_{fecha_str}.json'),
        os.path.join(BASE_DIR, 'driver_names.json'),
    ):
        if os.path.exists(dn_path):
            try:
                with open(dn_path, encoding='utf-8') as f:
                    local_dn = json.load(f)
                break
            except Exception:
                pass

    # 3. Mensajes desde PostgreSQL
    msgs = database.load_day_messages(fecha_str)
    print(f'[HIST] {fecha_str}: {len(msgs)} mensajes desde DB', flush=True)

    # 4. Procesar mensajes con estado local (sin tocar globales)
    local_bop_reports        = {}
    local_bo_responses       = {}
    local_last_bops_by_phone = {}

    def _local_add_media(bop, item):
        if bop and bop in local_bop_reports:
            local_bop_reports[bop].setdefault('media', []).append(item)
            if item['type'] == 'image':
                local_bop_reports[bop]['imgs'] = local_bop_reports[bop].get('imgs', 0) + 1

    for msg in msgs:
        phone  = msg.get('from', '')
        nombre = msg.get('from_name') or phone
        ts     = msg.get('timestamp', 0)
        mtype  = msg.get('type', '')
        hora   = fmt_hour(ts) if ts else ''

        if mtype == 'location':
            loc = msg.get('location', {})
            lat = loc.get('latitude') or loc.get('lat')
            lon = loc.get('longitude') or loc.get('lng')
            if lat and lon:
                item = {
                    'type': 'location', 'preview': loc.get('preview', ''),
                    'url': f'https://maps.google.com/?q={float(lat)},{float(lon)}',
                    'lat': float(lat), 'lon': float(lon),
                }
                for bop in local_last_bops_by_phone.get(phone, []):
                    _local_add_media(bop, item)
            continue

        if mtype == 'image':
            img     = msg.get('image') or {}
            caption = img.get('caption', '') or ''
            item = {
                'type': 'image', 'preview': img.get('preview', ''),
                'id': img.get('id', ''), 'caption': caption,
            }
            target_bops = extract_bop(caption) if caption else local_last_bops_by_phone.get(phone, [])
            for bop in target_bops:
                _local_add_media(bop, item)
            if not caption:
                continue
            text = caption

        # FIX: Soporte de videos en reconstrucción histórica (antes eran ignorados)
        elif mtype == 'video':
            vid     = msg.get('video') or {}
            caption = vid.get('caption', '') or ''
            item = {
                'type': 'video', 'preview': vid.get('preview', ''),
                'id': vid.get('id', ''), 'caption': caption,
            }
            target_bops = extract_bop(caption) if caption else local_last_bops_by_phone.get(phone, [])
            for bop in target_bops:
                _local_add_media(bop, item)
            if not caption:
                continue
            text = caption

        elif mtype == 'text':
            text = (msg.get('text') or {}).get('body', '')
            if not text:
                continue
        else:
            continue

        if phone in BO_PHONES or is_bo_fmt(text):
            r = parse_bo(text)
            if r:
                for bop in r['bops']:
                    if bop not in local_bo_responses:
                        local_bo_responses[bop] = {
                            'bo_status': r['bo_status'], 'bo_obs': r['bo_obs'],
                            'msgs': [], 'hora': hora,
                        }
                    else:
                        local_bo_responses[bop]['bo_status'] = r['bo_status']
                        local_bo_responses[bop]['bo_obs']    = r['bo_obs']
                        local_bo_responses[bop]['hora']      = hora
                    local_bo_responses[bop]['msgs'].append(f'{hora} {nombre}: {text[:80]}')
            continue

        if not is_driver_msg(text):
            continue
        r = parse_driver(text)
        if r:
            # ── TYPO CORRECTOR: corrección en reconstrucción histórica ────────
            # Construir indice punto para este dia historico si no existe
            if not hasattr(_build_day_payload_from_db, '_brp_cache'):
                _build_day_payload_from_db._brp_cache = {}
            _brp_local = _build_day_payload_from_db._brp_cache.get(fecha_str, {})

            _bops_hist, _corr_hist = corregir_lista_bops(
                r['bops'], r.get('ruta'), r.get('punto'),
                local_rutas, local_bop_to_ruta, _brp_local,
            )
            for _c in _corr_hist:
                print(
                    f'[TYPO-HIST] {_c["bop_original"]} → {_c["bop_corregido"]} '
                    f'(ruta={_c["ruta_real"]}, dist={_c["distancia"]})',
                    flush=True
                )
            # ─────────────────────────────────────────────────────────────────
            for bop in _bops_hist:
                ruta_real = local_bop_to_ruta.get(bop) or r['ruta'] or '?'
                if bop not in local_bop_reports:
                    local_bop_reports[bop] = {
                        'phone': phone, 'nombre': nombre, 'ruta': ruta_real,
                        'punto': r['punto'], 'status': r['status'], 'obs': r['obs'],
                        'ts': ts, 'hora': hora, 'msgs': [], 'imgs': 0, 'media': [],
                    }
                else:
                    local_bop_reports[bop]['status'] = r['status']
                    local_bop_reports[bop]['obs']    = r['obs']
                    local_bop_reports[bop]['hora']   = hora
                local_bop_reports[bop]['msgs'].append(f'{hora} {nombre}: {text[:100]}')
            local_last_bops_by_phone[phone] = list(_bops_hist)

    # 5. Construir con el helper compartido
    return _build_payload_from_state(
        local_bop_reports,
        local_bo_responses,
        local_rutas,
        local_bop_to_ruta,
        local_dn,
        fecha_str,
    )


@app.get('/api/history')
def api_history():
    """
    Devuelve todos los dias disponibles: historico (data.js o PostgreSQL) + hoy en vivo.
    - 18-23: se leen de data.js (generados antes de pasar a PostgreSQL)
    - 24 en adelante (sin incluir hoy): se reconstruyen desde PostgreSQL
    - Hoy: payload en vivo desde RAM
    """
    import re as _re

    # Fechas pre-PostgreSQL: leer data.js
    HIST_DIRS = [
        ('2026-03-18', 'dashboard_2026_03_18'),
        ('2026-03-19', 'dashboard_2026_03_19'),
        ('2026-03-20', 'dashboard_2026_03_20'),
        ('2026-03-21', 'dashboard_2026_03_21'),
        ('2026-03-23', 'dashboard_2026_03_23'),
    ]

    dias = []
    for fecha, dirname in HIST_DIRS:
        data_js = os.path.join(BASE_DIR, dirname, 'data.js')
        if not os.path.exists(data_js):
            continue
        try:
            with open(data_js, encoding='utf-8') as f:
                content = f.read()
            json_str = _re.sub(r'^const\s+dashboardData\s*=\s*', '', content).rstrip(';\n')
            day_data = json.loads(json_str)
            day_data['fecha'] = fecha
            dias.append(day_data)
        except Exception as e:
            print(f'[HIST] Error leyendo {data_js}: {e}')

    # Fechas desde PostgreSQL: desde el 24 hasta ayer (hoy se sirve desde RAM)
    from datetime import date as _date, timedelta as _td
    DB_START = _date(2026, 3, 24)
    ayer = _date.fromisoformat(today_str) - _td(days=1)
    DB_HIST_DATES = []
    d = DB_START
    while d <= ayer:
        DB_HIST_DATES.append(d.isoformat())
        d += _td(days=1)

    for fecha in DB_HIST_DATES:
        if fecha == today_str:
            continue  # hoy se sirve desde RAM
        try:
            payload = _build_day_payload_from_db(fecha)
            if payload['detalle_reportados'] or payload['rutas']:
                dias.append(payload)
            else:
                # Sin datos en DB — intentar data.js como fallback
                data_js = os.path.join(BASE_DIR, f'dashboard_{fecha.replace("-", "_")}', 'data.js')
                if os.path.exists(data_js):
                    with open(data_js, encoding='utf-8') as f:
                        content = f.read()
                    json_str = _re.sub(r'^const\s+dashboardData\s*=\s*', '', content).rstrip(';\n')
                    day_data = json.loads(json_str)
                    day_data['fecha'] = fecha
                    dias.append(day_data)
                    print(f'[HIST] {fecha}: DB vacia, usando data.js de fallback')
        except Exception as e:
            print(f'[HIST] Error reconstruyendo {fecha} desde DB: {e}')

    # Hoy en vivo
    dias.append(_build_dashboard_payload())

    return {'dias': dias}


# ── ENDPOINT: Re-fetch completo de un día desde Whapi ────────────────────────
@app.post('/api/refetch-day/{fecha}')
def refetch_day(fecha: str):
    """
    Re-descarga TODOS los mensajes de Whapi para una fecha y los re-procesa.
    Útil para recuperar mensajes que el webhook perdió (reinicios, videos, etc.)

    Flujo:
      1. Descarga todos los mensajes del día desde los 6 chats autorizados
      2. Guarda los nuevos en raw_messages (ON CONFLICT los ignora)
      3. Limpia el estado en RAM y recarga desde DB

    Uso: POST /api/refetch-day/2026-03-26
    """
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', fecha):
        raise HTTPException(400, 'Formato de fecha inválido — usar YYYY-MM-DD')
    if not TOKEN:
        raise HTTPException(503, 'Sin WHAPI_TOKEN configurado')

    with state_lock:
        fecha_hoy = today_str

    print(f'[REFETCH] Iniciando re-fetch de {fecha} desde Whapi...', flush=True)

    # 1. Descargar mensajes desde Whapi
    try:
        msgs_whapi = _descargar_mensajes_whapi(fecha)
    except Exception as e:
        raise HTTPException(500, f'Error descargando desde Whapi: {e}')

    print(f'[REFETCH] {len(msgs_whapi)} mensajes descargados de Whapi para {fecha}', flush=True)

    # 2. Guardar en DB (save_raw_message ignora duplicados por whapi_message_id)
    nuevos = 0
    for msg in msgs_whapi:
        try:
            result = database.save_raw_message(msg)
            if result:  # None = ya existia
                nuevos += 1
        except Exception as e:
            print(f'[REFETCH] Error guardando mensaje: {e}', flush=True)

    print(f'[REFETCH] {nuevos} mensajes nuevos guardados en DB', flush=True)

    # 3. Si es el día actual: limpiar RAM y recargar
    if fecha == fecha_hoy:
        _reset_estado_dia()
        load_xlsx(fecha)
        cargar_driver_names_desde_disco()
        msgs_db = []
        try:
            msgs_db = database.load_day_messages(fecha)
        except Exception as e:
            print(f'[REFETCH] Error cargando DB: {e}', flush=True)
        for m in msgs_db:
            procesar_mensaje(m)
        kpis, total = regenerar_dashboard()
        return {
            'ok': True, 'fecha': fecha,
            'msgs_whapi': len(msgs_whapi), 'msgs_nuevos_en_db': nuevos,
            'recargado_en_vivo': True, 'bops': total, 'kpis': kpis,
        }

    # Para días históricos: el siguiente /api/history los reconstruirá
    return {
        'ok': True, 'fecha': fecha,
        'msgs_whapi': len(msgs_whapi), 'msgs_nuevos_en_db': nuevos,
        'recargado_en_vivo': False,
        'siguiente_paso': f'Nuevos mensajes guardados. El histórico de {fecha} se reconstruye en /api/history.',
    }

@app.post('/webhook/whatsapp')
async def webhook(request: Request):
    try:
        data = await request.json()
    except Exception:
        return {'status': 'error', 'reason': 'invalid json'}

    messages = data.get('messages', [])
    if not messages and 'chat_id' in data:
        messages = [data]

    procesados = 0
    for msg in messages:
        chat_id = msg.get('chat_id', '')
        if chat_id not in AUTHORIZED_CHAT_IDS:
            continue
        threading.Thread(target=_procesar_y_actualizar, args=(msg,), daemon=True).start()
        procesados += 1

    return {'status': 'ok', 'procesados': procesados}

def _procesar_y_actualizar(msg):
    chat_id = msg.get('chat_id', '')

    # 1. Guardar mensaje crudo en PostgreSQL y obtener su ID
    raw_id = None
    try:
        raw_id = database.save_raw_message(msg)
    except Exception as e:
        print(f'[DB] save_raw_message: {e}', flush=True)

    # 2. Casos especiales: xlsx de rutas (cualquier remitente), imagen de asignacion de Roberto
    if msg.get('type') == 'document':
        doc    = msg.get('document') or {}
        fname  = doc.get('filename', '')
        doc_id = doc.get('id', '')
        if fname.lower().endswith('.xlsx') and doc_id:
            descargar_xlsx_doc(doc_id, fname)
        return

    if (msg.get('type') == 'image' and msg.get('from') == ROBERTO_PHONE
            and chat_id == XLSX_CHAT_ID):
        img_id = (msg.get('image') or {}).get('id', '')
        if img_id:
            threading.Thread(target=procesar_imagen_asignacion, args=(img_id,), daemon=True).start()
        return

    # 3. Parsear el mensaje y actualizar RAM
    parsed = procesar_mensaje(msg)

    # 4. Persistir datos normalizados en PostgreSQL
    if parsed and raw_id:
        try:
            if parsed['type'] == 'driver':
                r = parsed['parsed']
                for bop in parsed['bops']:
                    database.save_driver_report(raw_id, {
                        'id_bop':        bop,
                        'ruta':          bop_to_ruta.get(bop) or r.get('ruta') or '?',
                        'punto':         r.get('punto', '?'),
                        'estatus':       r.get('status', ''),
                        'observaciones': r.get('obs', ''),
                        'nombre':        parsed.get('nombre', ''),
                    })
            elif parsed['type'] == 'bo':
                r = parsed['parsed']
                for bop in parsed['bops']:
                    database.save_bo_closure(raw_id, {
                        'id_bop':         bop,
                        'codigo_cierre':  r.get('bo_status', ''),
                        'detalle':        r.get('bo_obs', ''),
                        'instrucciones':  '',
                    })
        except Exception as e:
            print(f'[DB] save parsed data: {e}', flush=True)

    # 5. Log
    try:
        kpis, total = regenerar_dashboard()
        ts          = msg.get('timestamp', 0)
        hora        = fmt_hour(ts) if ts else ''
        nombre      = msg.get('from_name') or msg.get('from', '?')
        nombre_safe = nombre.encode('ascii', errors='replace').decode('ascii')
        print(f'[RT] {hora} {nombre_safe} | BOPs={total} exit={kpis["exitosos"]} fall={kpis["fallidos"]}', flush=True)
    except Exception as e:
        print(f'[RT] Error: {e}', flush=True)

if __name__ == '__main__':
    uvicorn.run(app, host='0.0.0.0', port=8000)
   